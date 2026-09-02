from __future__ import annotations

import json
import os
import sqlite3
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib import error, parse, request


DEFAULT_ABSTRACT_API_BASE = "https://companyenrichment.abstractapi.com/v1/"


def normalize_company_payload(payload: Mapping[str, Any] | None) -> dict[str, Any]:
    if payload is None:
        return {
            "name": "",
            "domain": "",
            "website": "",
            "industry": "",
            "description": "",
            "employee_count": None,
            "city": "",
            "country": "",
            "year_founded": None,
            "company_size": "",
            "raw_response": "{}",
        }

    location = payload.get("location") or {}
    if isinstance(location, str):
        city = location
        country = ""
    elif isinstance(location, Mapping):
        city = location.get("city") or location.get("locality") or ""
        country = location.get("country") or ""
    else:
        city = ""
        country = ""

    company_size = payload.get("company_size") or payload.get("size") or ""
    employee_count = payload.get("employee_count") or payload.get("employees")
    if isinstance(employee_count, str):
        employee_count = int(employee_count.replace(",", "")) if employee_count.replace(",", "").isdigit() else None

    normalized = {
        "name": payload.get("name") or payload.get("company_name") or payload.get("company") or "",
        "domain": payload.get("domain") or payload.get("website") or "",
        "website": payload.get("website") or payload.get("domain") or "",
        "industry": payload.get("industry") or payload.get("category") or "",
        "description": payload.get("description") or "",
        "employee_count": employee_count,
        "city": city,
        "country": country,
        "year_founded": payload.get("year_founded") or payload.get("founded") or None,
        "company_size": company_size,
        "raw_response": json.dumps(payload, ensure_ascii=False),
    }

    if normalized["domain"] and not normalized["website"]:
        normalized["website"] = f"https://{normalized['domain']}" if not normalized["domain"].startswith("http") else normalized["domain"]
    if normalized["website"] and not normalized["domain"] and normalized["website"].startswith("http"):
        normalized["domain"] = normalized["website"].split("//", 1)[-1].split("/", 1)[0].replace("www.", "")

    return normalized


def extract_company_names_from_roster(sheet: Any) -> list[str]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    idx_by_name = {name: index for index, name in enumerate(header)}
    company_column = "FT employer name"
    if company_column not in idx_by_name:
        for candidate in ("Employer", "Company", "Company Name", "Organization"):
            if candidate in idx_by_name:
                company_column = candidate
                break

    results: list[str] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if not row:
            continue
        value = row[idx_by_name.get(company_column, -1)] if company_column in idx_by_name else None
        if value is None:
            continue
        company_name = str(value).strip()
        if not company_name or company_name in seen:
            continue
        seen.add(company_name)
        results.append(company_name)

    return results


def fetch_company_data(company_name: str, api_key: str, *, base_url: str = DEFAULT_ABSTRACT_API_BASE, timeout: int = 20) -> dict[str, Any]:
    if not api_key:
        raise ValueError("ABSTRACT_API_KEY is required")

    params = {
        "api_key": api_key,
        "company_name": company_name,
    }
    encoded = parse.urlencode(params)
    url = f"{base_url}?{encoded}"
    req = request.Request(url, headers={"Accept": "application/json"})

    try:
        with request.urlopen(req, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"AbstractAPI request failed for {company_name!r}: {exc.code} {body}") from exc
    except error.URLError as exc:
        raise RuntimeError(f"AbstractAPI connection failed for {company_name!r}: {exc.reason}") from exc

    if isinstance(payload, dict):
        return payload
    raise ValueError(f"Unexpected AbstractAPI response for {company_name!r}: {payload!r}")


def ensure_database(db_path: str | os.PathLike[str]) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL UNIQUE,
            domain TEXT,
            website TEXT,
            industry TEXT,
            description TEXT,
            employee_count INTEGER,
            city TEXT,
            country TEXT,
            year_founded INTEGER,
            company_size TEXT,
            raw_response TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.commit()
    return conn


def upsert_company(conn: sqlite3.Connection, company: Mapping[str, Any]) -> None:
    name = (company.get("name") or "").strip()
    if not name:
        return

    row = normalize_company_payload(company)
    payload = (
        row["name"],
        row["domain"],
        row["website"],
        row["industry"],
        row["description"],
        row["employee_count"],
        row["city"],
        row["country"],
        row["year_founded"],
        row["company_size"],
        row["raw_response"],
    )

    conn.execute(
        """
        INSERT INTO companies (
            company_name, domain, website, industry, description,
            employee_count, city, country, year_founded, company_size, raw_response
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(company_name) DO UPDATE SET
            domain = excluded.domain,
            website = excluded.website,
            industry = excluded.industry,
            description = excluded.description,
            employee_count = excluded.employee_count,
            city = excluded.city,
            country = excluded.country,
            year_founded = excluded.year_founded,
            company_size = excluded.company_size,
            raw_response = excluded.raw_response,
            updated_at = CURRENT_TIMESTAMP
        """,
        payload,
    )
    conn.commit()


def load_company_records_from_roster(roster_path: str | os.PathLike[str]) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(roster_path, read_only=True, data_only=True)
    sheet = workbook.active
    return extract_company_names_from_roster(sheet)


def enrich_companies_for_roster(
    roster_path: str | os.PathLike[str],
    db_path: str | os.PathLike[str],
    api_key: str,
) -> list[str]:
    names = load_company_records_from_roster(roster_path)
    conn = ensure_database(db_path)
    inserted: list[str] = []

    for company_name in names:
        payload = fetch_company_data(company_name, api_key)
        normalized = normalize_company_payload(payload)
        upsert_company(conn, normalized)
        inserted.append(normalized["name"] or company_name)

    conn.close()
    return inserted


def main() -> None:
    roster_path = os.environ.get("ROSTER_PATH", str(Path(__file__).resolve().parents[2] / "grad-roster-excel" / "grad_roster.xlsx"))
    db_path = os.environ.get("COMPANY_DB_PATH", str(Path(__file__).resolve().parents[2] / "database" / "companies.db"))
    api_key = os.environ.get("ABSTRACT_API_KEY")

    if not api_key:
        raise SystemExit("Set ABSTRACT_API_KEY before running this ETL.")

    inserted = enrich_companies_for_roster(roster_path, db_path, api_key)
    print(f"Inserted/updated {len(inserted)} companies into {db_path}")


if __name__ == "__main__":
    main()
