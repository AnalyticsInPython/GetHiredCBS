from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    "Graduating year",
    "First name",
    "Last name",
    "Status",
    "Email",
    "Cluster/J-term",
    "Undergrad institution",
    "Summer internship company",
    "FT employer name",
    "FT employer industry",
    "FT job title",
    "FT function",
    "City",
    "State",
    "Country",
    "Dual Degree",
]

STATUS_OPTIONS = [
    "Employed",
    "Not seeking",
    "Returning to sponsoring company or previous employer",
    "Starting a new business",
]


def build_workbook() -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(header) + 2, 12
        )

    status_col_letter = get_column_letter(COLUMNS.index("Status") + 1)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    dv.add(f"{status_col_letter}2:{status_col_letter}1000")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"

    return wb, ws


def main() -> None:
    wb, _ = build_workbook()
    wb.save("grad_roster.xlsx")
    print("Created grad_roster.xlsx")


if __name__ == "__main__":
    main()
